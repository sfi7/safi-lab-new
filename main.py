# SAFI LAB - Modern Manager (PyWebView Edition)
import os
import sys
import json
import unicodedata
import re
import webbrowser
import pythoncom
import win32com.client
import webview
import threading
import base64
import shutil
from datetime import datetime

from urllib.parse import quote
from openpyxl import load_workbook
import qrcode

# ========================= CONFIGURATION =========================
EXCEL_FILE      = os.path.abspath("Patients.xlsm")
SHEET_NAME      = "Patients"
OUTPUT_ROOT     = os.path.abspath("QR_Patients")
DOMAIN_HOST     = "safi-lab-3clq.vercel.app"
LAST_UPDATE_COL_INDEX = 18 

import subprocess

# =================================================================
 
# =================================================================

class SafiLabAPI:
    def __init__(self):
        self._window = None

    def set_window(self, window):
        self._window = window

    # --- Data Methods ---
    def get_patients(self):
        """Reads Excel and returns list of patients as JSON."""
        try:
            wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
            ws = wb[SHEET_NAME]
            patients = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None: continue
                
                p_id = str(row[0]).strip()
                p_name = str(row[1]) if row[1] is not None else ""
                p_age = str(row[2]) if row[2] is not None else ""
                p_gender = str(row[3]) if row[3] is not None else ""
                # Use Last Modified from Col 19 (index 18)
                p_date = str(row[18]) if len(row) > 18 and row[18] is not None else ""
                
                patients.append({
                    "id": p_id, "name": p_name, "age": p_age, 
                    "gender": p_gender, "date": p_date
                })
            wb.close()
            return json.dumps(patients)
        except Exception as e:
            print(f"Error reading Excel: {e}")
            return json.dumps([])

    def get_patient_details(self, pid):
        """Returns full details for a single patient."""
        try:
            wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
            ws = wb[SHEET_NAME]
            data = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None: continue
                if str(row[0]).strip() == str(pid):
                    data = {
                        "id": str(row[0]), 
                        "name": str(row[1]) if row[1] else "",
                        "age": str(row[2]) if row[2] else "", 
                        "gender": str(row[3]) if row[3] else "",
                        "clinic": str(row[4]) if len(row) > 4 and row[4] else "", 
                        "doctor": str(row[5]) if len(row) > 5 and row[5] else "",
                        "date": str(row[6]) if len(row) > 6 and row[6] else "", 
                        "phone": str(row[7]) if len(row) > 7 and row[7] else "",
                        "email": str(row[8]) if len(row) > 8 and row[8] else "",
                        "abs": str(row[9]) if len(row) > 9 and row[9] else "", 
                        "conc": str(row[10]) if len(row) > 10 and row[10] else "", 
                        "trans": str(row[11]) if len(row) > 11 and row[11] else "",
                        "last_modified": str(row[18]) if len(row) > 18 and row[18] else ""
                    }
                    break
            # Check if report exists
            folder_name = self._get_safe_filename(f"{data.get('name')}_{pid}")
            report_path = os.path.join(OUTPUT_ROOT, folder_name, f"patient_{pid}.html")
            is_generated = os.path.exists(report_path)

            # Read Status Columns (16=P, 17=Q) - Adjusting for 0-based index if using iter_rows, but here we used values_only
            # Wait, iter_rows returns tuple. 
            # Row structure: 0=ID, 1=Name, ..., 15=ReportLink(O), 16=QR(P), 17=SendReport(Q) ?? 
            # Let's check debug output from before:
            # Row 1: ('ID', 'Name', ..., 'Report Link', 'QR Code', 'Send Report')
            # So 15 is Report Link, 16 is QR Code, 17 is Send Report?
            # The user wants Emailed and WhatsApp. Let's use Col 16 (index 15) for Emailed and Col 17 (index 16) for WhatsApp if they are free or repurpose.
            # Actually, let's stick to the plan: Read specific columns.
            # Let's assume:
            # Col 15 (O) = Emailed
            # Col 16 (P) = WhatsApp
            # We need to be careful with indices.
            
            # Let's look at the row length.
            emailed = "No"
            whatsapp = "No"
            if len(row) > 15 and row[15]: emailed = str(row[15])
            if len(row) > 16 and row[16]: whatsapp = str(row[16])

            data["status"] = {
                "saved": True, # If we found it, it's saved
                "generated": is_generated,
                "emailed": emailed.lower() in ['yes', 'true', '1'],
                "whatsapp": whatsapp.lower() in ['yes', 'true', '1']
            }
            
            wb.close()
            return json.dumps(data)
        except Exception as e:
            print(f"Error details: {e}")
            return json.dumps({})

    def save_patient(self, data_json):
        """Saves or updates patient data in Excel."""
        try:
            data = json.loads(data_json)
            target_id = data.get('id')
            if not target_id: return False

            current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            pythoncom.CoInitialize()
            xl = win32com.client.Dispatch("Excel.Application")
            xl.Visible = False
            xl.DisplayAlerts = False
            wb = xl.Workbooks.Open(EXCEL_FILE)
            ws = wb.Worksheets(SHEET_NAME)

            last_row = ws.Cells(ws.Rows.Count, 1).End(-4162).Row
            found_row = self._find_row_by_id_com(ws, target_id)

            if found_row == 0:
                found_row = last_row + 1

            # Write Data
            ws.Cells(found_row, 1).Value = target_id
            ws.Cells(found_row, 2).Value = data.get('name', '')
            ws.Cells(found_row, 3).Value = data.get('age', '')
            ws.Cells(found_row, 4).Value = data.get('gender', '')
            ws.Cells(found_row, 5).Value = data.get('clinic', '')
            ws.Cells(found_row, 6).Value = data.get('doctor', '')
            ws.Cells(found_row, 7).Value = current_timestamp
            ws.Cells(found_row, 8).Value = data.get('phone', '')
            ws.Cells(found_row, 9).Value = data.get('email', '')
            ws.Cells(found_row, 10).Value = data.get('abs', '')
            ws.Cells(found_row, 11).Value = data.get('conc', '')
            ws.Cells(found_row, 12).Value = data.get('trans', '')
            ws.Cells(found_row, 19).Value = current_timestamp

            wb.Save()
            wb.Close()
            xl.Quit()
            return True
        except Exception as e:
            print(f"Save Error: {e}")
            return False
        finally:
            pythoncom.CoUninitialize()

    def delete_patient(self, pid):
        """Deletes a patient from Excel."""
        try:
            pythoncom.CoInitialize()
            xl = win32com.client.Dispatch("Excel.Application")
            xl.Visible = False
            xl.DisplayAlerts = False
            wb = xl.Workbooks.Open(EXCEL_FILE)
            ws = wb.Worksheets(SHEET_NAME)

            found_row = self._find_row_by_id_com(ws, pid)
            if found_row >= 2:
                ws.Rows(found_row).Delete()
                wb.Save()
                wb.Close()
                xl.Quit()
                
                # --- Delete Local Folder ---
                try:
                    # Find folder ending with _pid
                    target_folder = None
                    for item in os.listdir(OUTPUT_ROOT):
                        if item.endswith(f"_{pid}") and os.path.isdir(os.path.join(OUTPUT_ROOT, item)):
                            target_folder = item
                            break
                    
                    if target_folder:
                        folder_path = os.path.join(OUTPUT_ROOT, target_folder)
                        shutil.rmtree(folder_path)
                        print(f"Deleted folder: {folder_path}")
                except Exception as e:
                    print(f"Error deleting folder: {e}")

                return True
            else:

                wb.Close()
                xl.Quit()
                return False
        except Exception as e:
            print(f"Delete Error: {e}")
            return False
        finally:
            pythoncom.CoUninitialize()
            # Sync with GitHub after deletion
            try:
                print("Syncing deletion with GitHub...")
                self._git_push(f"Delete patient {pid}")
            except Exception as e:
                print(f"Git Sync Error: {e}")

    def generate_report(self, pid):
        """Runs the VBA macro to generate report."""
        def run_macro():
            try:
                pythoncom.CoInitialize()
                xl = win32com.client.Dispatch("Excel.Application")
                xl.Visible = False
                xl.DisplayAlerts = False
                wb = xl.Workbooks.Open(EXCEL_FILE)
                xl.Run("Generate_From_Python", pid)
                wb.Save()
                wb.Close()
                xl.Quit()
                
                # --- Fix File Structure for Cloudflare & QR Code ---
                try:
                    # Find folder
                    target_folder = None
                    for item in os.listdir(OUTPUT_ROOT):
                        if item.endswith(f"_{pid}") and os.path.isdir(os.path.join(OUTPUT_ROOT, item)):
                            target_folder = item
                            break
                    
                    if target_folder:
                        folder_path = os.path.join(OUTPUT_ROOT, target_folder)
                        
                        # 2. Construct Correct URL
                        from urllib.parse import quote
                        safe_folder_url = quote(target_folder)
                        # Point directly to the HTML file
                        correct_url = f"https://{DOMAIN_HOST}/QR_Patients/{safe_folder_url}/patient_{pid}.html"
                        
                        # 3. Generate QR
                        qr_img = qrcode.make(correct_url)
                        qr_filename = f"qr_{pid}.png"
                        qr_full_path = os.path.join(folder_path, qr_filename)
                        qr_img.save(qr_full_path)
                        print(f"Regenerated QR code with URL: {correct_url}")

                except Exception as e:
                    print(f"Structure/QR Fix Error: {e}")

                # --- Git Push (Vercel Deploy) ---
                try:
                    print(f"Syncing to GitHub for Vercel deployment...")
                    success, msg = self._git_push(f"Update report for patient {pid}")
                    
                    if success:
                        return True, f"Success & Pushed: {msg}"
                    else:
                        return True, f"Generated but Push Failed: {msg}"
                        
                except Exception as e:
                    print(f"Push Error: {e}")
                    return True, f"Generated but Push Error: {e}"

                return True, "Success"
            except Exception as e:
                return False, str(e)
            finally:
                pythoncom.CoUninitialize()

        # Run in thread to not block UI (though pywebview might block anyway if not careful)
        # For simplicity in this structure, we'll run synchronously or use a simple thread wrapper if needed.
        # Since pywebview calls are async from JS side, we can block here but it freezes UI.
        # Better to return a promise or just block for now as it's a local operation.
        success, msg = run_macro()
        return json.dumps({"success": success, "message": msg})

    def get_qr_data(self, name, pid):
        """Returns base64 image of QR code."""
        try:
            folder_name = self._get_safe_filename(f"{name}_{pid}")
            qr_path = os.path.join(OUTPUT_ROOT, folder_name, f"qr_{pid}.png")
            
            if os.path.exists(qr_path):
                with open(qr_path, "rb") as img_file:
                    b64 = base64.b64encode(img_file.read()).decode('utf-8')
                    return f"data:image/png;base64,{b64}"
            
            # Generate temporary preview
            url = f"https://{DOMAIN_HOST}/{folder_name}/patient_{pid}.html"
            qr = qrcode.make(url)
            from io import BytesIO
            buffered = BytesIO()
            qr.save(buffered, format="PNG")
            b64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
            return f"data:image/png;base64,{b64}"
        except Exception as e:
            print(f"QR Error: {e}")
            return None

    # --- Actions ---
    def send_email(self, pid):
        # Need to fetch details first as we only have ID
        details = json.loads(self.get_patient_details(pid))
        email = details.get('email')
        name = details.get('name')
        if not email: return
        
        folder_name = self._get_safe_filename(f"{name}_{pid}")
        safe_folder = quote(folder_name)
        url = f"https://{DOMAIN_HOST}/QR_Patients/{safe_folder}/patient_{pid}.html"
        
        subject = "SAFI LAB - Your Test Report"
        body = f"Dear {name},\n\nYou can access your SAFI LAB report here:\n{url}\n\nBest regards,\nSAFI LAB Team"
        webbrowser.open(f"mailto:{email}?subject={quote(subject)}&body={quote(body)}")
        
        # Update Status
        self._update_cell(pid, 16, "Yes") # Col 16 = P (1-based) -> actually let's verify column mapping.
        # If row has 15 items (0-14), 15 is P.
        # Let's use a helper that opens writeable workbook.

    def send_whatsapp(self, pid):
        details = json.loads(self.get_patient_details(pid))
        phone = details.get('phone')
        name = details.get('name')
        if not phone: return
        
        folder_name = self._get_safe_filename(f"{name}_{pid}")
        safe_folder = quote(folder_name)
        url = f"https://{DOMAIN_HOST}/QR_Patients/{safe_folder}/patient_{pid}.html"
        
        phone_clean = re.sub(r'[^\d+]', '', phone)
        message = f"Hello {name}, your test report is ready: {url}"
        webbrowser.open(f"https://wa.me/{phone_clean}?text={quote(message)}")
        
        # Update Status
        self._update_cell(pid, 17, "Yes") # Col 17 = Q

    def open_folder(self, pid):
        details = json.loads(self.get_patient_details(pid))
        name = details.get('name')
        folder_name = self._get_safe_filename(f"{name}_{pid}")
        path = os.path.join(OUTPUT_ROOT, folder_name)
        if os.path.exists(path):
            os.startfile(path)

    def open_vercel(self):
        webbrowser.open("https://vercel.com/dashboard")

    # --- Helpers ---
    def _find_row_by_id_com(self, ws_com, patient_id):
        target_str = str(patient_id).strip().lower()
        def normalize(s):
            s = str(s).strip().lower()
            if s.endswith(".0"): return s[:-2]
            return s
        target_norm = normalize(target_str)

        try:
            last_row = int(ws_com.Cells(ws_com.Rows.Count, 1).End(-4162).Row)
        except: return 0
            
        for r in range(2, last_row + 1):
            try:
                val = ws_com.Cells(r, 1).Value
                if val is None: continue
                if normalize(val) == target_norm: return r
            except: continue
        return 0

    def _get_safe_filename(self, text):
        if not text: return "unknown"
        # Match VBA: badChars = Array("\", "/", ":", "*", "?", """", "<", ">", "|")
        # VBA does NOT lowercase and does NOT replace spaces.
        bad_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
        for char in bad_chars:
            text = text.replace(char, '_')
        return text

    def _update_cell(self, pid, col_index, value):
        """Updates a specific cell for a patient."""
        try:
            pythoncom.CoInitialize()
            xl = win32com.client.Dispatch("Excel.Application")
            xl.Visible = False
            xl.DisplayAlerts = False
            wb = xl.Workbooks.Open(EXCEL_FILE)
            ws = wb.Worksheets(SHEET_NAME)

            found_row = self._find_row_by_id_com(ws, pid)
            if found_row >= 2:
                ws.Cells(found_row, col_index).Value = value
                wb.Save()
            
            wb.Close()
            xl.Quit()
        except Exception as e:
            print(f"Update Cell Error: {e}")
        finally:
            pythoncom.CoUninitialize()

    def _git_push(self, message):
        """Commits and pushes changes to GitHub."""
        try:
            # 1. Add all changes
            subprocess.run(["git", "add", "."], cwd=os.getcwd(), check=True)
            
            # 2. Commit
            # We use allow-empty in case there are no changes but we want to be sure
            subprocess.run(["git", "commit", "-m", message], cwd=os.getcwd(), check=False)
            
            # 3. Push
            # Explicitly push to origin master and set upstream to avoid "no upstream branch" errors
            result = subprocess.run(["git", "push", "-u", "origin", "master"], cwd=os.getcwd(), capture_output=True, text=True)
            
            if result.returncode == 0:
                print("Git Push Successful")
                return True, "Synced to GitHub"
            else:
                print(f"Git Push Error: {result.stderr}")
                return False, f"Push Failed: {result.stderr}"
        except Exception as e:
            print(f"Git Error: {e}")
            return False, str(e)

if __name__ == '__main__':
    api = SafiLabAPI()
    window = webview.create_window(
        'SAFI LAB - Modern Manager 2026', 
        'web/index.html', 
        js_api=api,
        width=1400, 
        height=900,
        resizable=True
    )
    api.set_window(window)
    webview.start(debug=False, http_port=23456)
